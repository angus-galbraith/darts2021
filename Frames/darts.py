from tkinter import *
import os
import openpyxl


root = Tk()
root.title("Darts")
root.geometry("450x375")

class statistics():

    def __init__(self):
        check_file = "darts.xlsx"
        titles = ["Name", "Sets", "Legs", "180's", "160+", "140+", "120+", "100+","80+", "60+", "Average",
         "Doubles Thrown", "Doubles Hit", "Checkout %"]

        if os.path.isfile(check_file):
            print("file exists")
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Results'
            col_num = 2
            for items in titles:
                sheet.cell(row= 2, column=col_num).value = items
                col_num += 1
            wb.save('darts.xlsx')

    def update_stats(self):
        wb = openpyxl.load_workbook('darts.xlsx')
        s = wb.active
        last_row = len(list(s.rows))
        last_row += 1
        col_num = 2
        values = list(screen.player1.stats.values())
        for value in values:
            s.cell(row=last_row, column=col_num).value = value
            col_num += 1
        last_row += 1  
        col_num = 2  
        values = list(screen.player2.stats.values())
        for value in values:
            s.cell(row=last_row, column=col_num).value = value
            col_num += 1
        
        wb.save('darts.xlsx')

            

class mainFrame(Frame):

    def __init__(self, parent=None, **options):
        LabelFrame.__init__(self, parent, **options)
        self.pack()

        # Create Menus 
        self.my_menu = Menu(root)
        root.config(menu=self.my_menu)
        self.options_menu= Menu(self.my_menu)
        self.my_menu.add_cascade(label='501', menu=self.options_menu) 
        self.options_menu.add_command(label='New Game', command=self.new_game_window)
        

        

    def add_frames(self):
       self.frame_one = Frame(self)
       self.frame_one.pack(side=LEFT, expand=YES, fill=BOTH)
       self.frame_three = Frame(self)
       self.frame_three.pack(side=LEFT, expand=YES, fill=BOTH)
       self.frame_two = Frame(self)
       self.frame_two.pack(side=LEFT, expand=YES, fill=BOTH)

    def new_game_window(self):
        self.win = Toplevel()
        Label(self.win, text="Player 1:").grid(row=0, column=0)
        self.pl1ent = Entry(self.win)
        self.pl1ent.grid(row=0,column=1)
        Label(self.win, text="Player 2:").grid(row=1, column=0)
        self.pl2ent = Entry(self.win)
        self.pl2ent.grid(row=1,column=1)
        Label(self.win, text="Sets: First to:-").grid(row=2, column=0)
        self.sets_spinbox = Spinbox(self.win, values=(1,2,3,4))
        self.sets_spinbox.grid(row=2, column=1)
        Label(self.win, text="Legs: First to:-").grid(row=3, column=0)
        self.legs_spinbox = Spinbox(self.win, values=(1,2,3,4))
        self.legs_spinbox.grid(row=3, column=1)
        Button(self.win, text="Start", command=self.start_game).grid(row=4, column=1,columnspan=2)

    def start_game(self):
        self.player1 = player()  #create 2 instances of the player class
        self.player2 = player()
        self.player1.stats["name"]= self.pl1ent.get()  # set the players names
        self.player2.stats["name"] = self.pl2ent.get()
        self.add_frames()
        self.screen_refresh(self.frame_one, self.frame_two, self.frame_three, self.player1.stats, self.player2.stats) # populate the three screens
        self.legs_to_win = int(self.legs_spinbox.get()) # set variables for legs and sets 
        self.sets_to_win = int(self.sets_spinbox.get())
        self.leg_to_play = 1  # starting variables for legs and sets
        self.set_to_play = 1
        self.win.destroy()
        self.to_throw()

    
    

    def screen_refresh(self, frame, frame1, frame2, player, player1):
        player = player
        frame = frame
        rownum = 0
        for (key, value) in player.items():
            Label(frame, text=key).grid(row=rownum, column=0)
            Label(frame, text=value).grid(row=rownum, column=1)
            rownum += 1
        player1 = player1
        frame1 = frame1
        rownum = 0
        for (key, value) in player1.items():
            Label(frame1, text=key).grid(row=rownum, column=0)
            Label(frame1, text=value).grid(row=rownum, column=1)
            rownum += 1
        frame2 = frame2
        Label(frame2, text=self.player1.stats["name"], font=(None, 25)).grid(row=0, column=0)
        Label(frame2, text=self.player2.stats["name"], font=(None, 25)).grid(row=0, column=1)
        self.pl1_remaining = Label(frame2, text=self.player1.score["remaining"], font=(None, 40), width=3, height=1)
        self.pl1_remaining.grid(row=1, column=0)
        self.pl1_remaining.configure(bg="white")
        self.pl2_remaining = Label(frame2, text=self.player2.score["remaining"], font=(None, 40), width=3, height=1)
        self.pl2_remaining.grid(row=1, column=1)
        self.pl2_remaining.configure(bg="white")
        self.score_ent = Entry(frame2, width=10)
        self.score_ent.grid(row=2,column=0)
        Button(frame2, text="Enter Score", command=self.button_pressed).grid(row=2, column=1)


    def to_throw(self): # called on al the start of a  leg
        self.player1.score['remaining'] = 501
        self.player2.score['remaining'] = 501
        self.screen_refresh(self.frame_one, self.frame_two, self.frame_three, self.player1.stats, self.player2.stats)
        if self.set_to_play % 2 != 0:
            if self.leg_to_play %2 != 0:
                self.current_player = 1
                self.player_to_throw()
            else:
                self.current_player = 2
                self.player_to_throw()
        else:
            if self.leg_to_play %2 != 0:
                self.current_player = 2
                self.player_to_throw()
            else:
                self.current_player = 1
                self.player_to_throw()


    def player_to_throw(self): # callled for each throw
        if self.current_player == 1:
            self.screen_refresh(self.frame_one, self.frame_two, self.frame_three, self.player1.stats, self.player2.stats)
            self.pl1_remaining.configure(bg="yellow")
            self.pl2_remaining.configure(bg="white")
            self.score_ent.focus()
        else:
            self.screen_refresh(self.frame_one, self.frame_two, self.frame_three, self.player1.stats, self.player2.stats)
            self.pl2_remaining.configure(bg="yellow")
            self.pl1_remaining.configure(bg="white")
            self.score_ent.focus()
        



    def button_pressed(self):
        score = int(self.score_ent.get())
        if self.current_player == 1:
            self.current_player = 2
            self.player1.score_entered(score)
            self.player_to_throw()

        else:
            self.current_player = 1
            self.player2.score_entered(score)
            self.player_to_throw()


class player:
    def __init__(self,):

        self.name = "Player"
        
        self.stats = {
            "name": "Player",
            "sets": 0,
            "legs": 0,
            "180": 0,
            "160": 0,
            "140": 0,
            "120": 0,
            "100": 0,
            "80": 0,
            "60": 0,
            "Average": 0,
            "Darts at double": 0,
            "Doubles Hit": 0,
            "Checkout %": 0,
            
        }

        self.score = {
            "remaining": 501,
            "totalscore": 0,
            "totaldarts": 0,
            
        }

    def score_entered(self, score):
        score = score
        self.score['remaining'] -= score
        self.score['totalscore'] += score
        if score == 180:
            self.stats['180'] += 1
        elif score >= 160:
            self.stats['160'] += 1
        elif score >= 140:
            self.stats['140'] += 1
        elif score >= 120:
            self.stats['120'] += 1
        elif score >= 100:
            self.stats['100'] += 1
        elif score >= 80:
            self.stats['80'] += 1
        elif score >= 60:
            self.stats['60'] += 1
        
        if self.score['remaining'] == 0:
            self.leg_won()
        elif self.score['remaining'] <= 50:
            self.darts_at_double()
        
        self.score['totaldarts'] += 3
        self.calculate_averages()

        
        return

    def darts_at_double(self):
        self.doubles_window = Toplevel()
        Label(self.doubles_window, text="Number of Darts at Doubles").grid(row=0, column=0)
        self.darts_doubles = Spinbox(self.doubles_window, values=(0,1,2,3))
        self.darts_doubles.grid(row=0, column=1)
        Button(self.doubles_window, text="Enter ", command=self.doubles_button_pressed).grid(row=1, column=1, columnspan=2)
        self.doubles_window.attributes('-topmost', 'true')
        

    def doubles_button_pressed(self):
        at_doubles = int(self.darts_doubles.get())
        self.stats['Darts at double'] += at_doubles
        self.doubles_window.destroy()
        screen.player_to_throw()

    def calculate_averages(self):
        self.stats['Average'] = round(3*(self.score['totalscore']/self.score['totaldarts']), 1)
        
    def leg_won(self):
        self.leg_window = Toplevel()
        Label(self.leg_window, text="Number of Darts Used").grid(row=0, column=0)
        self.darts_used = Spinbox(self.leg_window, values=(0,1,2,3))
        self.darts_used.grid(row=0, column=1)
        Label(self.leg_window, text="Number of Darts at Doubles").grid(row=1, column=0)
        self.darts_doubles = Spinbox(self.leg_window, values=(0,1,2,3))
        self.darts_doubles.grid(row=1, column=1)
        Button(self.leg_window, text='Enter', command=self.leg_button_pressed).grid(row=2, column=1, columnspan=2)
        self.leg_window.attributes('-topmost', 'true')

    def leg_button_pressed(self):
        self.stats['legs'] += 1
        self.stats['Doubles Hit'] += 1
        doubledarts = int(self.darts_doubles.get())
        self.stats['Darts at double'] += doubledarts
        self.stats["Checkout %"] = (self.stats['Doubles Hit']/self.stats['Darts at double'])*100
        screen.leg_to_play += 1
        totaldarts = int(self.darts_used.get())
        self.score['totaldarts'] += totaldarts
        self.leg_window.destroy()
        if self.stats['legs'] == screen.legs_to_win:
            self.set_won()
        else:
            screen.to_throw()
        

    def set_won(self):
        print('at sets')
        self.stats['sets'] += 1
        if self.stats['sets'] == screen.sets_to_win:
            self.game_won()
        else:
            screen.player1.stats['legs'] = 0
            screen.player2.stats['legs'] = 0
            screen.leg_to_play = 1
            screen.set_to_play += 1
            screen.to_throw()





    def game_won(self):
        print("Game over")
        player_stats.update_stats()




        


player_stats = statistics()
screen = mainFrame(root)
screen.mainloop()
